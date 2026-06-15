from __future__ import annotations

import math
import re
from dataclasses import dataclass
from pathlib import Path
from typing import BinaryIO

import pandas as pd
from openpyxl import load_workbook

from validators import normalize_spaces


EXCLUDED_SHEETS = {
    "Daily Activity w Contents",
    "Daily Activity w Contents (2)",
    "Daily Repair Rate",
    "Daily Repair Rate (Old)",
    "Daily Repair Rate (2)",
    "Daily Chart",
    "Sheet1",
}


@dataclass
class ProjectParseReport:
    parsed_rows: int = 0
    skipped_blocks: int = 0
    parsed_sheets: int = 0
    warnings: list[str] | None = None

    def __post_init__(self) -> None:
        if self.warnings is None:
            self.warnings = []


def _is_number(value: object) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool) and math.isfinite(float(value))


def _ratio_from_row(ws, row: int, col: int) -> float | None:
    for offset in range(1, 5):
        value = ws.cell(row, col + offset).value
        if _is_number(value):
            ratio = float(value)
            if 0 <= ratio <= 1:
                return ratio
    return None


def _parse_length_ft(value: object) -> float | None:
    text = normalize_spaces(value).lower()
    if not text:
        return None
    feet_inches = re.search(r"(\d+(?:\.\d+)?)\s*'\s*(\d+(?:\.\d+)?)?", text)
    if feet_inches:
        feet = float(feet_inches.group(1))
        inches = float(feet_inches.group(2) or 0)
        return feet + inches / 12
    number = re.search(r"(\d+(?:\.\d+)?)", text)
    if number and ("ft" in text or "feet" in text):
        return float(number.group(1))
    return None


def parse_project_pipe_repairs(file: str | Path | BinaryIO, report_date: object) -> tuple[pd.DataFrame, ProjectParseReport]:
    report = ProjectParseReport()
    try:
        wb = load_workbook(file, read_only=False, data_only=True)
    except Exception as exc:
        report.warnings.append(f"Project sheet'leri okunamadı: {exc}")
        return pd.DataFrame(), report

    date_value = pd.to_datetime(report_date, errors="coerce")
    if pd.isna(date_value):
        report.warnings.append("Project sheet parser: geçerli rapor tarihi yok.")
        return pd.DataFrame(), report
    date_text = date_value.strftime("%Y-%m-%d")

    rows: list[dict[str, object]] = []
    for sheet_name in wb.sheetnames:
        if sheet_name in EXCLUDED_SHEETS:
            continue
        ws = wb[sheet_name]
        sheet_rows = 0
        max_row = min(ws.max_row, 240)
        max_col = min(ws.max_column, 190)
        for row_idx in range(1, max_row + 1):
            # Project sheets are laid out as repeated 5-column pipe cards:
            # A:E, F:J, K:O, ... . Only scan card anchor columns to avoid
            # reading unrelated cells and to keep upload parsing responsive.
            for col_idx in range(1, max_col + 1, 5):
                cell = ws.cell(row_idx, col_idx)
                value = cell.value
                if not isinstance(value, str):
                    continue
                label = value.lower()
                if "repair" not in label or "rate" not in label:
                    continue

                ratio = _ratio_from_row(ws, row_idx, col_idx)
                pipe_length_ft = _parse_length_ft(ws.cell(row_idx + 4, col_idx).value)
                pipe_no = ws.cell(row_idx + 4, col_idx + 1).value
                repair_amount = ws.cell(row_idx + 5, col_idx + 1).value
                unit = normalize_spaces(ws.cell(row_idx + 5, col_idx + 2).value).lower().replace(".", "")

                if not (_is_number(pipe_no) and _is_number(repair_amount) and ratio is not None and unit == "m"):
                    report.skipped_blocks += 1
                    continue

                repair_count = ws.cell(cell.row + 5, cell.column + 3).value
                category = normalize_spaces(ws.cell(cell.row + 5, cell.column + 4).value)
                surface_state = normalize_spaces(ws.cell(cell.row + 1, cell.column).value)
                if "repair" in surface_state.lower():
                    surface_state = ""

                rows.append(
                    {
                        "date": date_text,
                        "project_sheet": normalize_spaces(sheet_name),
                        "block_cell": cell.coordinate,
                        "pipe_no": int(pipe_no),
                        "pipe_length_ft": pipe_length_ft,
                        "repair_amount": float(repair_amount),
                        "repair_ratio": float(ratio),
                        "repair_count": int(repair_count) if _is_number(repair_count) else None,
                        "repair_category": category or "Unspecified",
                        "surface_state": surface_state or "Unspecified",
                    }
                )
                sheet_rows += 1

        if sheet_rows:
            report.parsed_sheets += 1

    df = pd.DataFrame(rows)
    report.parsed_rows = len(df)
    if report.parsed_rows == 0:
        report.warnings.append("Project sheet'lerinden güvenli pipe-level repair kaydı çıkarılamadı.")
    return df, report
