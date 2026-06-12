from __future__ import annotations

from pathlib import Path
from typing import BinaryIO
from datetime import date

import pandas as pd
from openpyxl import load_workbook

from validators import (
    ValidationReport,
    coerce_number,
    coerce_report_date,
    normalize_spaces,
    normalize_status,
    validate_dataframe,
)


SHEET_CANDIDATES = ["Daily Repair Rate", "Daily Repair Rate (2)", "Daily Repair Rate (Old)"]

# Fixed workbook contract:
# - report date is in P1
# - old format source table is A3:P25, data rows 4:25
# - new format has two fixed source tables:
#   Plate: A3:P14, data rows 5:14
#   Coil: A15:P41, data rows 17:41
# Lower calculation/chart tables are never read.
REPORT_DATE_CELL = "P1"

OLD_FORMAT_SECTIONS = [
    {"production_type": "Coil", "start_row": 4, "end_row": 25},
]

NEW_FORMAT_SECTIONS = [
    {
        "production_type": "Plate",
        "title_cell": "A3",
        "title_text": "Ongoing & Recently Completed Productions with Plate",
        "start_row": 5,
        "end_row": 14,
    },
    {
        "production_type": "Coil",
        "title_cell": "A15",
        "title_text": "Ongoing & Recently Completed Productions with Coil",
        "start_row": 17,
        "end_row": 41,
    },
]

COLUMN_MAP = {
    "project_no": "B",
    "dimensions": "E",
    "qty": "H",
    "project_total_pipe_length": "I",
    "repaired_pipes_total_length": "J",
    "repaired_spiral_length": "K",
    "total_repair_amount": "L",
    "total_repair_amount_incl_skelp": "M",
    "project_status": "N",
    "repair_ratio": "O",
    "repair_ratio_incl_skelp": "P",
}


def _select_sheet(wb):
    available = [name for name in SHEET_CANDIDATES if name in wb.sheetnames]
    if not available:
        return None

    dated = []
    for name in available:
        report_date = coerce_report_date(wb[name][REPORT_DATE_CELL].value)
        dated.append((report_date, name))

    dated.sort(key=lambda item: (item[0] is not None, item[0] or date.min), reverse=True)
    return dated[0][1]


def _is_new_format(ws) -> bool:
    for section in NEW_FORMAT_SECTIONS:
        title = normalize_spaces(ws[section["title_cell"]].value).lower()
        expected = normalize_spaces(section["title_text"]).lower()
        if title != expected:
            return False
    return True


def _production_type(default_type: str, project_no: str) -> str:
    # Old-format reports did not have separate Plate/Coil blocks. Known plate
    # projects are still classified as Plate so trend filters stay consistent.
    lowered = project_no.lower()
    if "(pt)" in lowered or "plate" in lowered:
        return "Plate"
    return default_type


def parse_daily_repair_rate(file: str | Path | BinaryIO) -> tuple[pd.DataFrame, ValidationReport]:
    report = ValidationReport()

    try:
        wb = load_workbook(file, read_only=True, data_only=True)
    except Exception as exc:
        report.add_check("Sheet bulundu mu?", False)
        report.add_error(f"Excel dosyası açılamadı: {exc}")
        return pd.DataFrame(), report

    sheet_name = _select_sheet(wb)
    if sheet_name is None:
        report.add_check("Sheet bulundu mu?", False)
        report.add_error(f"Daily Repair Rate sheet'i bulunamadı. Beklenen sheet adları: {', '.join(SHEET_CANDIDATES)}")
        return pd.DataFrame(), report

    report.add_check("Sheet bulundu mu?", True)
    ws = wb[sheet_name]

    report_date = coerce_report_date(ws[REPORT_DATE_CELL].value)
    report.add_check("Tarih bulundu mu?", report_date is not None)
    if report_date is None:
        report.add_error(f"Rapor tarihi boş veya geçersiz: {sheet_name}!{REPORT_DATE_CELL}")
        return pd.DataFrame(), report

    sections = NEW_FORMAT_SECTIONS if _is_new_format(ws) else OLD_FORMAT_SECTIONS

    rows: list[dict[str, object]] = []
    for section in sections:
        for excel_row in range(section["start_row"], section["end_row"] + 1):
            project_no = normalize_spaces(ws[f"{COLUMN_MAP['project_no']}{excel_row}"].value)
            dimensions = normalize_spaces(ws[f"{COLUMN_MAP['dimensions']}{excel_row}"].value)

            if not project_no or not dimensions:
                continue

            row = {
                "date": report_date.isoformat(),
                "production_type": _production_type(section["production_type"], project_no),
                "project_no": project_no,
                "dimensions": dimensions,
                "project_status": normalize_status(ws[f"{COLUMN_MAP['project_status']}{excel_row}"].value),
                "excel_row": excel_row,
            }

            for col_name in [
                "qty",
                "project_total_pipe_length",
                "repaired_pipes_total_length",
                "repaired_spiral_length",
                "total_repair_amount",
                "total_repair_amount_incl_skelp",
                "repair_ratio",
                "repair_ratio_incl_skelp",
            ]:
                row[col_name] = coerce_number(ws[f"{COLUMN_MAP[col_name]}{excel_row}"].value)

            rows.append(row)

    df = pd.DataFrame(rows)
    if not df.empty:
        ordered = [
            "date",
            "production_type",
            "project_no",
            "dimensions",
            "qty",
            "project_total_pipe_length",
            "repaired_pipes_total_length",
            "repaired_spiral_length",
            "total_repair_amount",
            "total_repair_amount_incl_skelp",
            "project_status",
            "repair_ratio",
            "repair_ratio_incl_skelp",
            "excel_row",
        ]
        df = df[ordered]

    return df, validate_dataframe(df, report)
